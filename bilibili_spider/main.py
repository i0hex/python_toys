import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

base_url: str = "https://www.bilibili.com/v/popular/rank/"

sub_urls: list[dict[str, str]] = [
    # video
    {
        "all": "全部", "douga": "动画", "game": "游戏", "kichiku": "鬼畜", 
        "music": "音乐", "dance": "舞蹈", "cinephile": "影视", "ent": "娱乐", 
        "knowledge": "知识", "tech": "科技数码", "food": "美食", "car": "汽车", 
        "fashion": "时尚美妆", "sports": "体育运动", "animal": "动物"
    },
    # film
    {
        "anime": "番剧", "guochuang": "国创", "documentary": "纪录片", "movie": "电影", 
        "tv": "电视剧", "variety": "综艺"
    }
]

fetched_data: dict[str, list[list[str | None]]] = {}

sheet_headers: list[list[str]] = [
    # video header
    ["排名", "标题", "链接", "UP", "UP链接", "播放量", "弹幕量"],
    # film header
    ["排名", "标题", "链接", "更新状态", "播放量", "追番人数"]
]

def do_save(wb: Workbook, index: int, sub_type: int, sub_url: str, sub_data: list[list[str | None]]) -> None:

    max_width_each_column: list[float] = [0] * len(sheet_headers[sub_type])

    ws: Worksheet = wb.create_sheet(sub_urls[sub_type][sub_url], index)

    # sheet header
    ws.append(sheet_headers[sub_type])
    header_font: Font = Font(bold=True, size=12)
    for cell in ws[1]:
        cell.font = header_font

    # maximum column width
    min_width = 8
    # minimum column width
    max_width = 50
    # font scale factor
    font_factor = 1.1

    for data in sub_data:
        ws.append(data)
        # compare the maximum width of each column
        for i, v in enumerate(data):
            v_len = len(v) if v is not None else 0
            final_width = max(min_width, min(max_width, v_len * font_factor + 2))
            if final_width > max_width_each_column[i]:
                max_width_each_column[i] = final_width

    # setting the appropriate width of each column
    for column_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(column_index)
        ws.column_dimensions[column_letter].width = max_width_each_column[column_index - 1]

    print(f"Save data from ['{sub_urls[sub_type][sub_url]}']({base_url + sub_url}).")


def save() -> None:
    if len(fetched_data) <= 0:
        print("Err: No data to save.")
        return

    print("Saving data...")

    wb: Workbook | None = None
    
    for index, (sub_url, sub_data) in enumerate(fetched_data.items()):
        if wb is None:
            wb = Workbook()

        if sub_url in sub_urls[0].keys():
            if len(sub_data) > 0:
                do_save(wb, index, 0, sub_url, sub_data)
        elif sub_url in sub_urls[1].keys():
            if len(sub_data) > 0:
                do_save(wb, index, 1, sub_url, sub_data)

    timpstamp: int = int(time.time())
    if wb is not None:
        file_name: str = "bilibili_rank_data_" + str(timpstamp) + ".xlsx"
        wb.save(file_name)
        print(f"The Bilibili ranking data has been saved. File: {file_name}")
    else:
        print("Err: No data has been saved.")


def fetch(driver: WebDriver, sub_url: str) -> None:
    sub_type: int = -1
    if sub_url in sub_urls[0].keys():
        sub_type = 0
    elif sub_url in sub_urls[1].keys():
        sub_type = 1

    if sub_type == -1:
        print(f"Err: invalid url: {sub_url}")
        return

    url = base_url + sub_url
    # Implicit wait for 5 seconds
    driver.implicitly_wait(5)
    driver.get(url)

    fetched_data[sub_url] = []

    print(f"Fetch: ['{sub_urls[sub_type][sub_url]}']({url})")

    rank_items: list[WebElement] = driver.find_elements(By.CLASS_NAME, "rank-item")
    for item in rank_items:
        idx: str | None = item.get_attribute("data-rank")

        # elem[@class="info"]
        info: WebElement = item.find_element(By.CLASS_NAME, "info")

        # elem[@class="title"]
        info_title: WebElement = info.find_element(By.CLASS_NAME, "title")
        title: str | None = info_title.get_attribute("title")
        link: str | None = info_title.get_attribute("href")

        # elem[@class="detail"]
        info_detail: WebElement = info.find_element(By.CLASS_NAME, "detail")
        # elem[@class="detail-state"]
        info_detail_state: WebElement = info_detail.find_element(By.CLASS_NAME, "detail-state")
        # elem[@class="data-box"]
        info_detail_state_data: list[WebElement] = info_detail_state.find_elements(By.CLASS_NAME, "data-box")

        if sub_type == 0:
            # elem: <a>
            up_info: WebElement = info_detail.find_element(By.TAG_NAME, "a")
            fetched_data[sub_url].append([idx, title, link, up_info.text, up_info.get_attribute("href"), 
                                          info_detail_state_data[0].text, info_detail_state_data[1].text])
        elif sub_type == 1:
            # elem: <span>
            update_info: WebElement = info_detail.find_element(By.TAG_NAME, "span")
            fetched_data[sub_url].append([idx, title, link, update_info.text, 
                                          info_detail_state_data[0].text, info_detail_state_data[1].text])
    print(f"Fetch {len(rank_items)} records from ['{sub_urls[sub_type][sub_url]}']({url}).")

    
def spider() -> None:

    driver: WebDriver | None = None

    try:
        print("Welcome to use bilibili ranking spider:")
        print("0 - video")
        print("1 - film")
        print("2 - all")
        print("3 - quit")
        sub_type: int = int(input("Please enter the type you want to fetch: "))

        if sub_type < 0 or sub_type > 3:
            print("Err: invalid type.")
            return
        
        if sub_type == 3:
            exit(0)
        
        options: Options = Options()
        options.set_preference("general.useragent.override", "Mozilla/5.0 (Windows NT 10.0; Win64; x64)\
                AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        # options.set_preference("intl.accept_languages", "en-US")
        options.add_argument("--headless")

        service: Service = Service(executable_path='./geckodriver')
        driver = webdriver.Firefox(service=service, options=options)

        if sub_type == 2:
            for v in sub_urls:
                for sub_url in v.keys():
                    fetch(driver, sub_url)
        else:
            for sub_url in sub_urls[sub_type].keys():
                fetch(driver, sub_url)

        save()
        print("Done!")
    except Exception as e:
        print(f"Err: {str(e)}")
    finally:
        if driver is not None:
            driver.quit()
            print("Browser was closed!")
        print("Bye!")

if __name__ == "__main__":
    spider()
